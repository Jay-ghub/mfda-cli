"""
Excel (.xlsx) Reader Plugin

- **Format**: Handles `.xlsx` spreadsheets (OpenXML format) plain or compressed(gz, gzip).
- **Options**:
  - `sheet` (name or index, default: first sheet)
  - `header_row` (default: first row)
  - `limit` (optional row limit)
  - empty cells -> NULL
- **Errors Raised**:
  - `FileFormatError` if workbook is invalid or unreadable
  - `ConfigurationError` for sheet/index issues
- **Returns**: a `Table` (see core model docs)
"""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from mfda.errors import FileFormatError

NULL = None


class _Table:
    def __init__(self, columns: list[str], rows: list[list[Any]]) -> None:
        self.columns = columns
        self._rows = rows

    @property
    def shape(self) -> tuple[int, int]:
        return (len(self._rows), len(self.columns))

    def as_records(self) -> list[dict[str, Any]]:
        return [dict(zip(self.columns, row, strict=False)) for row in self._rows]


def read(
    path: str | Path,
    *,
    sheet: int | str | None = None,
    header_row: int = 0,
    limit: int | None = None,
) -> Any:
    wb = load_workbook(path, read_only=True)

    if sheet is None:
        ws = wb.active
    elif isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    elif isinstance(sheet, str):
        if sheet not in wb.sheetnames:
            raise FileFormatError(f"No such sheet: {sheet}")
        ws = wb[sheet]

    rows_iter = ws.iter_rows(values_only=True)

    # skip header_row
    for _ in range(header_row):
        next(rows_iter)

    # next row the header and force to string
    header = next(rows_iter)
    columns = [str(val) for val in header]

    # read rows data
    records: list[list[str | None]] = []
    for row in rows_iter:
        # Apply limit
        if limit is not None and len(records) >= limit:
            break
        clean_row = [NULL if val is None or val == "" else val for val in row]
        records.append(clean_row)

    # return table class
    return _Table(columns, records)
